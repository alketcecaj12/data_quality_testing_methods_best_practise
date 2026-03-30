"""
Data Validation Pipeline
=========================
Validates any tabular dataset across six data quality dimensions:

    1. Completeness  — are required fields present and populated?
    2. Validity      — do values conform to expected formats and domains?
    3. Uniqueness    — are key fields free of duplicates?
    4. Consistency   — are relationships between fields logically coherent?
    5. Timeliness    — are dates within expected ranges?
    6. Accuracy      — are numeric values within plausible bounds?

Outputs:
    - JSON summary  (machine-readable, for downstream pipelines)
    - Excel report  (human-readable, formatted workbook)
    - HTML report   (shareable, browser-viewable)

Usage:
    python validator.py --input transactions.csv --config config.yaml --output report

    # or with defaults:
    python validator.py --input transactions.csv
"""

import argparse
import json
import os
import time
import warnings
from dataclasses import dataclass, field, asdict
from datetime import datetime
from typing import Any

from logger import (
    get_logger, log_pipeline_start, log_pipeline_end,
    log_check_result, log_dimension_summary, log_warnings,
)

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")


# ── DATA STRUCTURES ───────────────────────────────────────────────────────────

@dataclass
class CheckResult:
    dimension:   str
    check_name:  str
    column:      str
    status:      str        # PASS | WARN | FAIL
    score:       float      # 0–100
    n_issues:    int
    n_total:     int
    pct_issues:  float
    detail:      str
    sample_bad:  list = field(default_factory=list)


@dataclass
class ValidationReport:
    dataset_name:    str
    run_timestamp:   str
    n_rows:          int
    n_columns:       int
    overall_score:   float
    overall_status:  str
    dimension_scores: dict
    checks:          list


# ── VALIDATION ENGINE ─────────────────────────────────────────────────────────

class DataValidator:

    # Default config — override by passing config dict
    DEFAULT_CONFIG = {
        "required_columns": [
            "transaction_id", "transaction_date", "amount",
            "currency", "transaction_type", "status",
            "sender_account", "receiver_account",
        ],
        "unique_columns":    ["transaction_id"],
        "not_null_columns":  ["transaction_id", "transaction_date", "amount", "currency"],
        "numeric_columns":   {"amount": {"min": 0, "max": 1e8}},
        "categorical_columns": {
            "currency":         ["DKK", "EUR", "USD", "GBP", "SEK", "NOK"],
            "transaction_type": ["PAYMENT", "TRANSFER", "REFUND", "FEE", "ADJUSTMENT"],
            "status":           ["COMPLETED", "PENDING", "FAILED", "REVERSED"],
            "country_code":     ["DK", "SE", "NO", "FI", "DE", "GB", "US", "FR"],
        },
        "date_columns":      ["transaction_date", "value_date", "created_at"],
        "date_range": {
            "transaction_date": {"min": "2020-01-01", "max": "today"},
            "value_date":       {"min": "2020-01-01", "max": "today+7"},
        },
        "consistency_rules": [
            {
                "name":   "value_date >= transaction_date",
                "col_a":  "value_date",
                "col_b":  "transaction_date",
                "op":     ">=",
            }
        ],
        "warn_threshold":    0.5,   # % issues → WARN
        "fail_threshold":    2.0,   # % issues → FAIL
    }

    def __init__(self, df: pd.DataFrame, config: dict = None, dataset_name: str = "dataset"):
        self.df = df.copy()
        self.cfg = {**self.DEFAULT_CONFIG, **(config or {})}
        self.dataset_name = dataset_name
        self.results: list[CheckResult] = []
        self.log = get_logger()

    # ── helpers ───────────────────────────────────────────────────────────────

    def _status(self, pct: float) -> str:
        if pct <= self.cfg["warn_threshold"]:
            return "PASS"
        elif pct <= self.cfg["fail_threshold"]:
            return "WARN"
        return "FAIL"

    def _score(self, pct: float) -> float:
        return round(max(0, 100 - pct * 2), 1)

    def _sample(self, series: pd.Series, n: int = 5) -> list:
        vals = series.dropna().unique()[:n]
        return [str(v) for v in vals]

    def _add(self, dimension, check_name, column, bad_mask, detail):
        n_total = len(self.df)
        n_issues = int(bad_mask.sum())
        pct = round(n_issues / n_total * 100, 2) if n_total else 0
        self.results.append(CheckResult(
            dimension=dimension,
            check_name=check_name,
            column=column,
            status=self._status(pct),
            score=self._score(pct),
            n_issues=n_issues,
            n_total=n_total,
            pct_issues=pct,
            detail=detail,
            sample_bad=self._sample(self.df.loc[bad_mask, column]) if column in self.df else [],
        ))
        log_check_result(
            self.log, dimension, check_name, column,
            self._status(pct), n_issues, n_total, pct, detail
        )

    # ── 1. COMPLETENESS ───────────────────────────────────────────────────────

    def check_completeness(self):
        # Required columns present
        missing_cols = [c for c in self.cfg["required_columns"] if c not in self.df.columns]
        mask = pd.Series([len(missing_cols) > 0] * len(self.df))
        self.results.append(CheckResult(
            dimension="Completeness",
            check_name="Required columns present",
            column="(schema)",
            status="FAIL" if missing_cols else "PASS",
            score=0.0 if missing_cols else 100.0,
            n_issues=len(missing_cols),
            n_total=len(self.cfg["required_columns"]),
            pct_issues=round(len(missing_cols) / max(len(self.cfg["required_columns"]), 1) * 100, 1),
            detail=f"Missing: {missing_cols}" if missing_cols else "All required columns present",
            sample_bad=missing_cols,
        ))

        # Null checks per column
        for col in self.cfg["not_null_columns"]:
            if col not in self.df.columns:
                continue
            mask = self.df[col].isna()
            self._add("Completeness", f"No nulls in {col}", col, mask,
                      f"{mask.sum()} null values in '{col}'")

        # Overall row completeness (any null in required cols)
        req_present = [c for c in self.cfg["required_columns"] if c in self.df.columns]
        if req_present:
            any_null = self.df[req_present].isna().any(axis=1)
            self._add("Completeness", "Row completeness (required fields)", "(row)",
                      any_null, f"{any_null.sum()} rows have at least one null in required fields")

    # ── 2. VALIDITY ───────────────────────────────────────────────────────────

    def check_validity(self):
        # Categorical domain checks
        for col, allowed in self.cfg["categorical_columns"].items():
            if col not in self.df.columns:
                continue
            mask = ~self.df[col].isna() & ~self.df[col].isin(allowed)
            self._add("Validity", f"Domain check: {col}", col, mask,
                      f"Values not in allowed set {allowed}")

        # Numeric range checks
        for col, bounds in self.cfg["numeric_columns"].items():
            if col not in self.df.columns:
                continue
            numeric = pd.to_numeric(self.df[col], errors="coerce")
            out_of_range = ~numeric.isna() & (
                (numeric < bounds.get("min", -np.inf)) |
                (numeric > bounds.get("max", np.inf))
            )
            self._add("Validity", f"Range check: {col}", col, out_of_range,
                      f"Values outside [{bounds.get('min')}, {bounds.get('max')}]")

        # Negative amount check (separate explicit rule)
        if "amount" in self.df.columns:
            neg = pd.to_numeric(self.df["amount"], errors="coerce") < 0
            self._add("Validity", "No negative amounts", "amount", neg,
                      f"{neg.sum()} negative amount values detected")

        # Date format checks
        for col in self.cfg["date_columns"]:
            if col not in self.df.columns:
                continue
            def is_bad_date(v):
                if pd.isna(v):
                    return False
                try:
                    pd.to_datetime(v)
                    return False
                except Exception:
                    return True
            bad = self.df[col].apply(is_bad_date)
            self._add("Validity", f"Date format: {col}", col, bad,
                      f"{bad.sum()} unparseable date values in '{col}'")

    # ── 3. UNIQUENESS ─────────────────────────────────────────────────────────

    def check_uniqueness(self):
        for col in self.cfg["unique_columns"]:
            if col not in self.df.columns:
                continue
            dupes = self.df.duplicated(subset=[col], keep=False)
            self._add("Uniqueness", f"Unique values: {col}", col, dupes,
                      f"{dupes.sum()} rows involved in duplicate {col} values")

    # ── 4. CONSISTENCY ────────────────────────────────────────────────────────

    def check_consistency(self):
        for rule in self.cfg["consistency_rules"]:
            col_a = rule["col_a"]
            col_b = rule["col_b"]
            op    = rule["op"]
            name  = rule["name"]
            if col_a not in self.df.columns or col_b not in self.df.columns:
                continue
            a = pd.to_datetime(self.df[col_a], errors="coerce")
            b = pd.to_datetime(self.df[col_b], errors="coerce")
            both_valid = ~a.isna() & ~b.isna()
            if op == ">=":
                violation = both_valid & (a < b)
            elif op == "<=":
                violation = both_valid & (a > b)
            elif op == "==":
                violation = both_valid & (a != b)
            else:
                violation = pd.Series([False] * len(self.df))
            self._add("Consistency", f"Consistency: {name}", col_a, violation,
                      f"{violation.sum()} rows where {name} is violated")

    # ── 5. TIMELINESS ─────────────────────────────────────────────────────────

    def check_timeliness(self):
        today = datetime.today()
        for col, bounds in self.cfg["date_range"].items():
            if col not in self.df.columns:
                continue
            dates = pd.to_datetime(self.df[col], errors="coerce")
            min_d = pd.to_datetime(bounds.get("min", "2000-01-01"))
            max_str = bounds.get("max", "today")
            if max_str == "today":
                max_d = pd.Timestamp(today)
            elif max_str.startswith("today+"):
                days = int(max_str.split("+")[1])
                max_d = pd.Timestamp(today) + pd.Timedelta(days=days)
            else:
                max_d = pd.to_datetime(max_str)

            out = ~dates.isna() & ((dates < min_d) | (dates > max_d))
            self._add("Timeliness", f"Date range: {col}", col, out,
                      f"{out.sum()} dates outside expected range [{min_d.date()} → {max_d.date()}]")

        # Stale data check — latest record vs today
        if "transaction_date" in self.df.columns:
            latest = pd.to_datetime(
                self.df["transaction_date"], errors="coerce"
            ).max()
            days_stale = (pd.Timestamp(today) - latest).days if pd.notna(latest) else 999
            stale = days_stale > 30
            self.results.append(CheckResult(
                dimension="Timeliness",
                check_name="Data freshness",
                column="transaction_date",
                status="WARN" if stale else "PASS",
                score=50.0 if stale else 100.0,
                n_issues=1 if stale else 0,
                n_total=1,
                pct_issues=100.0 if stale else 0.0,
                detail=f"Latest record is {days_stale} days old (latest: {latest})",
                sample_bad=[str(latest)] if stale else [],
            ))

    # ── 6. ACCURACY ───────────────────────────────────────────────────────────

    def check_accuracy(self):
        for col in self.cfg["numeric_columns"]:
            if col not in self.df.columns:
                continue
            numeric = pd.to_numeric(self.df[col], errors="coerce").dropna()
            if len(numeric) < 10:
                continue
            mean, std = numeric.mean(), numeric.std()
            extreme = ~pd.to_numeric(self.df[col], errors="coerce").isna() & (
                (pd.to_numeric(self.df[col], errors="coerce") - mean).abs() > 4 * std
            )
            self._add("Accuracy", f"Statistical outliers (>4σ): {col}", col, extreme,
                      f"{extreme.sum()} values more than 4 standard deviations from mean "
                      f"(mean={mean:,.0f}, std={std:,.0f})")

    # ── RUN ALL ───────────────────────────────────────────────────────────────

    def run(self) -> ValidationReport:
        self.log.info(f"Running validation checks on '{self.dataset_name}' ...")
        self.check_completeness()
        self.check_validity()
        self.check_uniqueness()
        self.check_consistency()
        self.check_timeliness()
        self.check_accuracy()

        # Dimension scores
        dim_scores = {}
        for dim in ["Completeness", "Validity", "Uniqueness",
                    "Consistency", "Timeliness", "Accuracy"]:
            dim_checks = [r for r in self.results if r.dimension == dim]
            dim_scores[dim] = round(
                np.mean([r.score for r in dim_checks]) if dim_checks else 100.0, 1
            )

        overall = round(np.mean(list(dim_scores.values())), 1)
        if overall >= 90:
            status = "PASS"
        elif overall >= 75:
            status = "WARN"
        else:
            status = "FAIL"

        return ValidationReport(
            dataset_name=self.dataset_name,
            run_timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            n_rows=len(self.df),
            n_columns=len(self.df.columns),
            overall_score=overall,
            overall_status=status,
            dimension_scores=dim_scores,
            checks=self.results,
        )


# ── JSON EXPORT ───────────────────────────────────────────────────────────────

def export_json(report: ValidationReport, path: str):
    data = asdict(report)
    with open(path, "w") as f:
        json.dump(data, f, indent=2, default=str)
    print(f"  ✓ JSON  → {path}")


# ── EXCEL EXPORT ──────────────────────────────────────────────────────────────

STATUS_COLORS = {"PASS": "00B050", "WARN": "FF9900", "FAIL": "C00000"}
STATUS_TEXT   = {"PASS": "FFFFFF", "WARN": "FFFFFF", "FAIL": "FFFFFF"}

def _hdr(ws, row, ncols, text, fill="1F4E79", size=11):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=size)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def _col_hdr(ws, row, headers, fill="2E75B6"):
    fill_obj = PatternFill("solid", fgColor=fill)
    font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = fill_obj
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _status_cell(cell, status):
    cell.fill = PatternFill("solid", fgColor=STATUS_COLORS.get(status, "808080"))
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    cell.alignment = Alignment(horizontal="center")

def export_excel(report: ValidationReport, path: str):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Dashboard ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Dashboard"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 40

    _hdr(ws, 1, 4, "  Data Validation Report", size=13)

    meta = [
        ("Dataset",        report.dataset_name),
        ("Run timestamp",  report.run_timestamp),
        ("Rows",           f"{report.n_rows:,}"),
        ("Columns",        str(report.n_columns)),
        ("Overall score",  f"{report.overall_score} / 100"),
        ("Overall status", report.overall_status),
    ]
    for r, (label, value) in enumerate(meta, start=3):
        ws.cell(r, 1, label).font = Font(bold=True, name="Arial", size=10)
        cell = ws.cell(r, 2, value)
        cell.font = Font(name="Arial", size=10)
        if label == "Overall status":
            _status_cell(cell, value)
        if label == "Overall score":
            score = report.overall_score
            cell.fill = PatternFill("solid", fgColor=(
                "00B050" if score >= 90 else "FF9900" if score >= 75 else "C00000"
            ))
            cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    ws.cell(10, 1, "Dimension Scores").font = Font(bold=True, name="Arial", size=10)
    _col_hdr(ws, 11, ["Dimension", "Score", "Status", "Note"])
    for r, (dim, score) in enumerate(report.dimension_scores.items(), start=12):
        status = "PASS" if score >= 90 else "WARN" if score >= 75 else "FAIL"
        ws.cell(r, 1, dim).font = Font(name="Arial", size=9)
        ws.cell(r, 2, score).font = Font(name="Arial", size=9)
        _status_cell(ws.cell(r, 3, status), status)
        ws.cell(r, 4, "").font = Font(name="Arial", size=9)
        for c in range(1, 5):
            ws.cell(r, c).border = _thin()

    # ── Sheet 2: All Checks ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("All Checks")
    col_widths = [16, 35, 18, 10, 12, 10, 10, 12, 45, 35]
    col_names  = ["Dimension", "Check", "Column", "Status", "Score",
                  "Issues", "Total", "Issue %", "Detail", "Sample Bad Values"]
    for i, (w, h) in enumerate(zip(col_widths, col_names), 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    _col_hdr(ws2, 1, col_names)
    ws2.row_dimensions[1].height = 25

    for r, chk in enumerate(report.checks, start=2):
        row_data = [
            chk.dimension, chk.check_name, chk.column,
            chk.status, chk.score,
            chk.n_issues, chk.n_total, f"{chk.pct_issues:.1f}%",
            chk.detail, ", ".join(chk.sample_bad),
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(r, c, val)
            cell.font = Font(name="Arial", size=9)
            cell.border = _thin()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if c == 4:
                _status_cell(cell, val)

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(col_names))}1"

    # ── Sheet 3: Failures Only ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Failures & Warnings")
    _col_hdr(ws3, 1, col_names)
    for i, w in enumerate(col_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    r = 2
    for chk in report.checks:
        if chk.status in ("FAIL", "WARN"):
            row_data = [
                chk.dimension, chk.check_name, chk.column,
                chk.status, chk.score, chk.n_issues, chk.n_total,
                f"{chk.pct_issues:.1f}%", chk.detail, ", ".join(chk.sample_bad),
            ]
            for c, val in enumerate(row_data, 1):
                cell = ws3.cell(r, c, val)
                cell.font = Font(name="Arial", size=9)
                cell.border = _thin()
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if c == 4:
                    _status_cell(cell, val)
            r += 1

    ws3.freeze_panes = "A2"

    wb.save(path)
    print(f"  ✓ Excel → {path}")


# ── HTML EXPORT ───────────────────────────────────────────────────────────────

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Data Validation Report — {{ report.dataset_name }}</title>
<style>
  body { font-family: Arial, sans-serif; font-size: 13px; color: #333;
         max-width: 1100px; margin: 40px auto; padding: 0 24px; background: #f8f8f8; }
  h1   { color: #1F4E79; font-size: 22px; margin-bottom: 4px; }
  h2   { color: #2E75B6; font-size: 15px; margin: 28px 0 10px; }
  .meta { color: #666; font-size: 12px; margin-bottom: 24px; }
  .score-banner { display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 28px; }
  .score-card { background: #fff; border: 1px solid #ddd; border-radius: 8px;
                padding: 14px 20px; min-width: 140px; text-align: center; }
  .score-card .label { font-size: 11px; color: #888; margin-bottom: 4px; }
  .score-card .value { font-size: 22px; font-weight: bold; }
  .PASS { color: #00B050; } .WARN { color: #FF9900; } .FAIL { color: #C00000; }
  table { width: 100%; border-collapse: collapse; background: #fff;
          border-radius: 8px; overflow: hidden; margin-bottom: 24px; }
  th { background: #1F4E79; color: #fff; padding: 8px 12px;
       text-align: left; font-size: 11px; font-weight: bold; }
  td { padding: 7px 12px; border-bottom: 1px solid #eee; font-size: 12px;
       vertical-align: top; }
  tr:last-child td { border-bottom: none; }
  tr:hover td { background: #f0f6ff; }
  .badge { display: inline-block; padding: 2px 10px; border-radius: 12px;
           font-size: 11px; font-weight: bold; color: #fff; }
  .badge-PASS { background: #00B050; }
  .badge-WARN { background: #FF9900; }
  .badge-FAIL { background: #C00000; }
  .footer { color: #aaa; font-size: 11px; margin-top: 40px; text-align: center; }
</style>
</head>
<body>
<h1>Data Validation Report</h1>
<div class="meta">
  Dataset: <b>{{ report.dataset_name }}</b> &nbsp;|&nbsp;
  Run: <b>{{ report.run_timestamp }}</b> &nbsp;|&nbsp;
  Rows: <b>{{ "{:,}".format(report.n_rows) }}</b> &nbsp;|&nbsp;
  Columns: <b>{{ report.n_columns }}</b>
</div>

<div class="score-banner">
  <div class="score-card">
    <div class="label">Overall Score</div>
    <div class="value {{ report.overall_status }}">{{ report.overall_score }}</div>
  </div>
  <div class="score-card">
    <div class="label">Overall Status</div>
    <div class="value {{ report.overall_status }}">{{ report.overall_status }}</div>
  </div>
  {% for dim, score in report.dimension_scores.items() %}
  <div class="score-card">
    <div class="label">{{ dim }}</div>
    {% set s = "PASS" if score >= 90 else "WARN" if score >= 75 else "FAIL" %}
    <div class="value {{ s }}">{{ score }}</div>
  </div>
  {% endfor %}
</div>

<h2>Failures &amp; Warnings</h2>
<table>
  <thead>
    <tr>
      <th>Dimension</th><th>Check</th><th>Column</th>
      <th>Status</th><th>Score</th><th>Issues</th><th>Issue %</th><th>Detail</th>
    </tr>
  </thead>
  <tbody>
    {% for chk in report.checks %}
    {% if chk.status in ["FAIL", "WARN"] %}
    <tr>
      <td>{{ chk.dimension }}</td>
      <td>{{ chk.check_name }}</td>
      <td><code>{{ chk.column }}</code></td>
      <td><span class="badge badge-{{ chk.status }}">{{ chk.status }}</span></td>
      <td>{{ chk.score }}</td>
      <td>{{ "{:,}".format(chk.n_issues) }}</td>
      <td>{{ chk.pct_issues }}%</td>
      <td>{{ chk.detail }}</td>
    </tr>
    {% endif %}
    {% endfor %}
  </tbody>
</table>

<h2>All Checks</h2>
<table>
  <thead>
    <tr>
      <th>Dimension</th><th>Check</th><th>Column</th>
      <th>Status</th><th>Score</th><th>Issues</th><th>Issue %</th>
    </tr>
  </thead>
  <tbody>
    {% for chk in report.checks %}
    <tr>
      <td>{{ chk.dimension }}</td>
      <td>{{ chk.check_name }}</td>
      <td><code>{{ chk.column }}</code></td>
      <td><span class="badge badge-{{ chk.status }}">{{ chk.status }}</span></td>
      <td>{{ chk.score }}</td>
      <td>{{ "{:,}".format(chk.n_issues) }}</td>
      <td>{{ chk.pct_issues }}%</td>
    </tr>
    {% endfor %}
  </tbody>
</table>

<div class="footer">Generated automatically by data_validator.py</div>
</body>
</html>
"""

def export_html(report: ValidationReport, path: str):
    from jinja2 import Template
    html = Template(HTML_TEMPLATE).render(report=report)
    with open(path, "w") as f:
        f.write(html)
    print(f"  ✓ HTML  → {path}")


# ── MAIN ──────────────────────────────────────────────────────────────────────

def run(input_path: str, output_prefix: str, dataset_name: str = None, config: dict = None):
    log = get_logger()
    t0  = time.time()

    df   = pd.read_csv(input_path)
    name = dataset_name or os.path.basename(input_path).replace(".csv", "")

    log_pipeline_start(log, "Data Validation", input_path, len(df), len(df.columns))

    validator = DataValidator(df, config=config, dataset_name=name)
    report    = validator.run()

    n_fail = sum(1 for r in report.checks if r.status == "FAIL")
    n_warn = sum(1 for r in report.checks if r.status == "WARN")
    n_pass = sum(1 for r in report.checks if r.status == "PASS")
    log.info(f"Checks complete  │  {len(report.checks)} total: {n_pass} PASS / {n_warn} WARN / {n_fail} FAIL")

    log_dimension_summary(log, report.dimension_scores)
    log.info(f"Overall score    │  {report.overall_score} / 100  [{report.overall_status}]")

    log.info("Exporting outputs ...")
    os.makedirs(os.path.dirname(output_prefix) or ".", exist_ok=True)
    export_json(report,  f"{output_prefix}.json")
    export_excel(report, f"{output_prefix}.xlsx")
    export_html(report,  f"{output_prefix}.html")

    fails = [r for r in report.checks if r.status == "FAIL"]
    if fails:
        log.error(f"Top failures ({len(fails)}):")
        for r in sorted(fails, key=lambda x: -x.pct_issues)[:5]:
            log.error(f"  • [{r.dimension}] {r.check_name} — {r.pct_issues}% ({r.n_issues:,} rows)")

    log_pipeline_end(log, "Data Validation", time.time() - t0, os.path.dirname(output_prefix) or ".")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Data Validation Pipeline")
    parser.add_argument("--input",   default="transactions.csv",  help="Input CSV path")
    parser.add_argument("--output",  default="validation_report", help="Output prefix (no extension)")
    parser.add_argument("--name",    default=None,                help="Dataset display name")
    args = parser.parse_args()
    run(args.input, args.output, args.name)