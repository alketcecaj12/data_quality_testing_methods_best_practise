"""
Data Profiling Pipeline
========================
Analyses the structure, content, and relationships within any tabular dataset.

Produces per-column and dataset-level profiles covering:
    - Schema & data types
    - Missing value analysis
    - Numeric distributions (mean, std, percentiles, skewness, kurtosis, outliers)
    - Categorical distributions (cardinality, top values, frequency)
    - Date distributions (min, max, range, gaps)
    - Correlation analysis (numeric columns)
    - Duplicate detection
    - Cross-column relationship hints

Outputs:
    - JSON  (machine-readable full profile)
    - Excel (human-readable workbook with per-column sheets)
    - HTML  (shareable browser report with inline charts)

Usage:
    python data_profiler.py --input transactions.csv --output profile_report
"""

import argparse
import base64
import io
import json
import os
import time
import warnings
from dataclasses import dataclass, field, asdict
from datetime import datetime
from typing import Any

from logger import (
    get_logger, log_pipeline_start, log_pipeline_end,
    log_profile_column, log_warnings,
)

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import scipy.stats as stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
plt.rcParams.update({
    "figure.facecolor": "white",
    "axes.facecolor":   "white",
    "axes.edgecolor":   "#cccccc",
    "axes.labelcolor":  "#333333",
    "xtick.color":      "#555555",
    "ytick.color":      "#555555",
    "font.family":      "sans-serif",
    "font.size":        9,
})


# ── DATA STRUCTURES ───────────────────────────────────────────────────────────

@dataclass
class ColumnProfile:
    name:           str
    dtype:          str
    inferred_type:  str        # numeric | categorical | datetime | boolean | text | mixed
    n_total:        int
    n_missing:      int
    pct_missing:    float
    n_unique:       int
    pct_unique:     float
    # numeric only
    mean:           float = None
    std:            float = None
    min:            float = None
    p25:            float = None
    p50:            float = None
    p75:            float = None
    p95:            float = None
    max:            float = None
    skewness:       float = None
    kurtosis:       float = None
    n_zeros:        int   = None
    n_negative:     int   = None
    n_outliers_iqr: int   = None
    # categorical only
    top_values:     list  = field(default_factory=list)   # [(value, count, pct), ...]
    # datetime only
    date_min:       str   = None
    date_max:       str   = None
    date_range_days: int  = None
    # text only
    avg_length:     float = None
    max_length:     int   = None
    # chart (base64 png)
    chart_b64:      str   = None


@dataclass
class DatasetProfile:
    dataset_name:    str
    run_timestamp:   str
    source_path:     str
    n_rows:          int
    n_columns:       int
    n_duplicates:    int
    pct_duplicates:  float
    memory_mb:       float
    overall_completeness: float
    columns:         list   = field(default_factory=list)   # list[ColumnProfile]
    correlations:    dict   = field(default_factory=dict)
    high_correlations: list = field(default_factory=list)   # [(col_a, col_b, r), ...]
    warnings:        list   = field(default_factory=list)


# ── CHARTING ──────────────────────────────────────────────────────────────────

def _fig_to_b64(fig) -> str:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=100, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("utf-8")


def _chart_numeric(series: pd.Series, col_name: str) -> str:
    clean = series.dropna()
    if len(clean) < 5:
        return None
    fig, axes = plt.subplots(1, 2, figsize=(7, 2.5))
    fig.suptitle(col_name, fontsize=10, color="#1F4E79", fontweight="bold")

    # Histogram
    axes[0].hist(clean, bins=min(40, len(clean.unique())),
                 color="#2E75B6", edgecolor="white", linewidth=0.4, alpha=0.85)
    axes[0].set_title("Distribution", fontsize=8)
    axes[0].set_ylabel("Count", fontsize=7)

    # Box plot
    bp = axes[1].boxplot(clean, vert=True, patch_artist=True,
                         widths=0.5,
                         boxprops=dict(facecolor="#BDD7EE", color="#2E75B6"),
                         medianprops=dict(color="#C00000", linewidth=1.5),
                         whiskerprops=dict(color="#2E75B6"),
                         capprops=dict(color="#2E75B6"),
                         flierprops=dict(marker="o", color="#FF9900",
                                         markerfacecolor="#FF9900", markersize=3))
    axes[1].set_title("Box Plot", fontsize=8)
    axes[1].set_xticks([])

    plt.tight_layout()
    return _fig_to_b64(fig)


def _chart_categorical(value_counts: pd.Series, col_name: str) -> str:
    top = value_counts.head(10)
    fig, ax = plt.subplots(figsize=(7, 2.5))
    fig.suptitle(col_name, fontsize=10, color="#1F4E79", fontweight="bold")
    colors = ["#2E75B6"] + ["#BDD7EE"] * (len(top) - 1)
    bars = ax.barh(top.index.astype(str)[::-1], top.values[::-1],
                   color=colors[::-1], edgecolor="white", linewidth=0.4)
    ax.set_title("Top values by frequency", fontsize=8)
    ax.set_xlabel("Count", fontsize=7)
    for bar, val in zip(bars, top.values[::-1]):
        ax.text(bar.get_width() + max(top.values) * 0.01, bar.get_y() + bar.get_height() / 2,
                f"{val:,}", va="center", fontsize=7, color="#333")
    plt.tight_layout()
    return _fig_to_b64(fig)


def _chart_missing(df: pd.DataFrame) -> str:
    missing = df.isna().mean().sort_values(ascending=False)
    missing = missing[missing > 0]
    if missing.empty:
        return None
    fig, ax = plt.subplots(figsize=(8, max(2, len(missing) * 0.35)))
    ax.barh(missing.index[::-1], missing.values[::-1] * 100,
            color="#C00000", alpha=0.75, edgecolor="white")
    ax.set_xlabel("% Missing", fontsize=8)
    ax.set_title("Missing values by column", fontsize=10,
                 color="#1F4E79", fontweight="bold")
    ax.set_xlim(0, 100)
    for i, v in enumerate(missing.values[::-1]):
        ax.text(v * 100 + 0.5, i, f"{v*100:.1f}%", va="center", fontsize=7)
    plt.tight_layout()
    return _fig_to_b64(fig)


def _chart_correlation(corr_matrix: pd.DataFrame) -> str:
    if corr_matrix.empty or len(corr_matrix) < 2:
        return None
    fig, ax = plt.subplots(figsize=(max(4, len(corr_matrix) * 0.8),
                                    max(3, len(corr_matrix) * 0.7)))
    im = ax.imshow(corr_matrix.values, cmap="RdYlGn", vmin=-1, vmax=1, aspect="auto")
    ax.set_xticks(range(len(corr_matrix.columns)))
    ax.set_yticks(range(len(corr_matrix.index)))
    ax.set_xticklabels(corr_matrix.columns, rotation=45, ha="right", fontsize=7)
    ax.set_yticklabels(corr_matrix.index, fontsize=7)
    for i in range(len(corr_matrix)):
        for j in range(len(corr_matrix.columns)):
            val = corr_matrix.values[i, j]
            ax.text(j, i, f"{val:.2f}", ha="center", va="center",
                    fontsize=7, color="black" if abs(val) < 0.7 else "white")
    plt.colorbar(im, ax=ax, fraction=0.046)
    ax.set_title("Correlation matrix (numeric columns)", fontsize=9,
                 color="#1F4E79", fontweight="bold")
    plt.tight_layout()
    return _fig_to_b64(fig)


# ── PROFILING ENGINE ──────────────────────────────────────────────────────────

def _infer_type(series: pd.Series) -> str:
    if series.dtype == bool or set(series.dropna().unique()) <= {0, 1, True, False, "0", "1"}:
        return "boolean"
    if pd.api.types.is_numeric_dtype(series):
        return "numeric"
    if pd.api.types.is_datetime64_any_dtype(series):
        return "datetime"
    # Try parsing as date
    sample = series.dropna().head(50)
    try:
        pd.to_datetime(sample, infer_datetime_format=True)
        return "datetime"
    except Exception:
        pass
    # Text vs categorical
    n_unique = series.nunique()
    n_total  = series.count()
    if n_total > 0 and n_unique / n_total < 0.5 and n_unique <= 50:
        return "categorical"
    return "text"


def profile_column(series: pd.Series) -> ColumnProfile:
    n_total   = len(series)
    n_missing = int(series.isna().sum())
    n_unique  = int(series.nunique(dropna=True))
    inferred  = _infer_type(series)

    prof = ColumnProfile(
        name          = series.name,
        dtype         = str(series.dtype),
        inferred_type = inferred,
        n_total       = n_total,
        n_missing     = n_missing,
        pct_missing   = round(n_missing / n_total * 100, 2) if n_total else 0,
        n_unique      = n_unique,
        pct_unique    = round(n_unique / (n_total - n_missing) * 100, 2) if (n_total - n_missing) > 0 else 0,
    )

    clean = series.dropna()

    if inferred == "numeric":
        numeric = pd.to_numeric(clean, errors="coerce").dropna()
        if len(numeric) > 0:
            q1, q3  = numeric.quantile(0.25), numeric.quantile(0.75)
            iqr     = q3 - q1
            prof.mean           = round(float(numeric.mean()), 4)
            prof.std            = round(float(numeric.std()), 4)
            prof.min            = round(float(numeric.min()), 4)
            prof.p25            = round(float(q1), 4)
            prof.p50            = round(float(numeric.median()), 4)
            prof.p75            = round(float(q3), 4)
            prof.p95            = round(float(numeric.quantile(0.95)), 4)
            prof.max            = round(float(numeric.max()), 4)
            prof.skewness       = round(float(stats.skew(numeric)), 4)
            prof.kurtosis       = round(float(stats.kurtosis(numeric)), 4)
            prof.n_zeros        = int((numeric == 0).sum())
            prof.n_negative     = int((numeric < 0).sum())
            prof.n_outliers_iqr = int(((numeric < q1 - 1.5 * iqr) | (numeric > q3 + 1.5 * iqr)).sum())
            prof.chart_b64      = _chart_numeric(numeric, series.name)

    elif inferred == "categorical":
        vc = clean.value_counts()
        prof.top_values = [
            (str(v), int(c), round(c / len(clean) * 100, 1))
            for v, c in vc.head(15).items()
        ]
        prof.chart_b64 = _chart_categorical(vc, series.name)

    elif inferred == "datetime":
        dates = pd.to_datetime(clean, errors="coerce").dropna()
        if len(dates) > 0:
            prof.date_min        = str(dates.min().date())
            prof.date_max        = str(dates.max().date())
            prof.date_range_days = (dates.max() - dates.min()).days

    elif inferred == "text":
        lengths = clean.astype(str).apply(len)
        prof.avg_length = round(float(lengths.mean()), 1)
        prof.max_length = int(lengths.max())
        vc = clean.value_counts()
        prof.top_values = [
            (str(v), int(c), round(c / len(clean) * 100, 1))
            for v, c in vc.head(10).items()
        ]
        prof.chart_b64 = _chart_categorical(vc, series.name)

    return prof


def profile_dataset(df: pd.DataFrame,
                    dataset_name: str = "dataset",
                    source_path: str  = "") -> DatasetProfile:
    log = get_logger()

    ds = DatasetProfile(
        dataset_name         = dataset_name,
        run_timestamp        = datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        source_path          = source_path,
        n_rows               = len(df),
        n_columns            = len(df.columns),
        n_duplicates         = int(df.duplicated().sum()),
        pct_duplicates       = round(df.duplicated().sum() / max(len(df), 1) * 100, 2),
        memory_mb            = round(df.memory_usage(deep=True).sum() / 1024 / 1024, 3),
        overall_completeness = round((1 - df.isna().mean().mean()) * 100, 2),
    )

    log.info(f"Profiling {len(df.columns)} columns ...")
    for col in df.columns:
        cp = profile_column(df[col])
        ds.columns.append(cp)
        extra = ""
        if cp.inferred_type == "numeric" and cp.mean is not None:
            extra = f"mean={cp.mean:,.2f}  outliers={cp.n_outliers_iqr}"
        elif cp.inferred_type == "categorical" and cp.top_values:
            extra = f"top={cp.top_values[0][0]} ({cp.top_values[0][2]}%)"
        elif cp.inferred_type == "datetime":
            extra = f"{cp.date_min} → {cp.date_max}"
        log_profile_column(log, col, cp.inferred_type, cp.pct_missing, cp.n_unique, extra)

    # Correlation matrix (numeric columns only)
    num_cols = [c for c in df.columns
                if pd.api.types.is_numeric_dtype(df[c]) and df[c].nunique() > 2]
    if len(num_cols) >= 2:
        corr = df[num_cols].corr().round(3)
        ds.correlations     = corr.to_dict()
        ds.high_correlations = [
            (col_a, col_b, round(float(corr.loc[col_a, col_b]), 3))
            for i, col_a in enumerate(num_cols)
            for col_b in num_cols[i+1:]
            if abs(corr.loc[col_a, col_b]) >= 0.7
        ]

    # Dataset-level warnings
    if ds.pct_duplicates > 1:
        ds.warnings.append(f"High duplicate rate: {ds.pct_duplicates:.1f}% ({ds.n_duplicates:,} rows)")
    if ds.overall_completeness < 95:
        ds.warnings.append(f"Low overall completeness: {ds.overall_completeness:.1f}%")
    for cp in ds.columns:
        if cp.pct_missing > 10:
            ds.warnings.append(f"'{cp.name}' has {cp.pct_missing:.1f}% missing values")
        if cp.inferred_type == "numeric" and cp.n_outliers_iqr and cp.n_outliers_iqr > 0:
            pct = round(cp.n_outliers_iqr / cp.n_total * 100, 1)
            if pct > 5:
                ds.warnings.append(f"'{cp.name}' has {pct}% IQR outliers ({cp.n_outliers_iqr:,} rows)")
        if cp.pct_unique == 100 and cp.n_total > 100:
            ds.warnings.append(f"'{cp.name}' appears to be a unique identifier (100% unique)")

    log_warnings(log, ds.warnings)

    if ds.high_correlations:
        log.info(f"High correlations detected ({len(ds.high_correlations)}):")
        for a, b, r in ds.high_correlations:
            log.info(f"  • {a} ↔ {b}  r={r}")

    return ds


# ── JSON EXPORT ───────────────────────────────────────────────────────────────

def export_json(ds: DatasetProfile, path: str):
    data = asdict(ds)
    # Remove chart blobs from JSON — they're large and not useful there
    for col in data.get("columns", []):
        col.pop("chart_b64", None)
    with open(path, "w") as f:
        json.dump(data, f, indent=2, default=str)
    print(f"  ✓ JSON  → {path}")


# ── EXCEL EXPORT ──────────────────────────────────────────────────────────────

def _thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_cell(ws, row, col, value, fill="1F4E79", size=10, bold=True):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=fill)
    c.font = Font(bold=bold, color="FFFFFF", name="Arial", size=size)
    c.alignment = Alignment(horizontal="left", vertical="center")
    return c

def _val_cell(ws, row, col, value, bold=False, color=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=9, bold=bold,
                  color=color if color else "000000")
    c.border = _thin()
    c.alignment = Alignment(vertical="top")
    return c

def export_excel(ds: DatasetProfile, path: str):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Overview ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Overview"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18

    _hdr_cell(ws, 1, 1, "  Data Profile Report", size=13)
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 28

    meta = [
        ("Dataset",            ds.dataset_name),
        ("Source",             ds.source_path),
        ("Run timestamp",      ds.run_timestamp),
        ("Rows",               f"{ds.n_rows:,}"),
        ("Columns",            str(ds.n_columns)),
        ("Duplicate rows",     f"{ds.n_duplicates:,} ({ds.pct_duplicates:.1f}%)"),
        ("Memory",             f"{ds.memory_mb:.2f} MB"),
        ("Overall completeness", f"{ds.overall_completeness:.1f}%"),
    ]
    for r, (label, value) in enumerate(meta, start=3):
        ws.cell(r, 1, label).font = Font(bold=True, name="Arial", size=10)
        ws.cell(r, 2, value).font = Font(name="Arial", size=10)

    # Warnings
    if ds.warnings:
        ws.cell(12, 1, "Warnings").font = Font(bold=True, name="Arial", size=10, color="C00000")
        for i, w in enumerate(ds.warnings, start=13):
            ws.cell(i, 1, f"⚠  {w}").font = Font(name="Arial", size=9, color="C00000")
            ws.merge_cells(f"A{i}:H{i}")

    # Column summary table
    start_row = 13 + len(ds.warnings) + 2
    ws.cell(start_row, 1, "Column Summary").font = Font(bold=True, name="Arial",
                                                         size=11, color="1F4E79")
    ws.merge_cells(f"A{start_row}:H{start_row}")

    hdr_row = start_row + 1
    hdrs = ["Column", "Type", "Inferred", "% Missing", "Unique", "% Unique", "Notes"]
    for c, h in enumerate(hdrs, 1):
        _hdr_cell(ws, hdr_row, c, h, fill="2E75B6", size=9)

    for r, cp in enumerate(ds.columns, start=hdr_row + 1):
        notes = ""
        if cp.inferred_type == "numeric":
            notes = f"mean={cp.mean:,.2f}, std={cp.std:,.2f}" if cp.mean is not None else ""
        elif cp.inferred_type == "categorical" and cp.top_values:
            notes = f"top: {cp.top_values[0][0]} ({cp.top_values[0][2]}%)"
        elif cp.inferred_type == "datetime":
            notes = f"{cp.date_min} → {cp.date_max}"

        row_data = [cp.name, cp.dtype, cp.inferred_type,
                    f"{cp.pct_missing:.1f}%", cp.n_unique, f"{cp.pct_unique:.1f}%", notes]
        for c, val in enumerate(row_data, 1):
            cell = _val_cell(ws, r, c, val)
            if c == 4 and cp.pct_missing > 10:
                cell.font = Font(name="Arial", size=9, color="C00000", bold=True)
            if c == 3:
                type_colors = {
                    "numeric": "DAEEF3", "categorical": "E2EFDA",
                    "datetime": "FFF2CC", "text": "F2DCDB", "boolean": "EBF0FB",
                }
                cell.fill = PatternFill("solid", fgColor=type_colors.get(cp.inferred_type, "F2F2F2"))

    # ── Sheet 2: Numeric Profiles ─────────────────────────────────────────────
    ws_num = wb.create_sheet("Numeric Profiles")
    num_cols_prof = [cp for cp in ds.columns if cp.inferred_type == "numeric"]
    if num_cols_prof:
        hdrs = ["Column", "Mean", "Std", "Min", "P25", "Median",
                "P75", "P95", "Max", "Skew", "Kurtosis",
                "Zeros", "Negative", "IQR Outliers", "% Missing"]
        widths = [18,12,12,12,12,12,12,12,12,10,10,10,10,14,12]
        for i, (h, w) in enumerate(zip(hdrs, widths), 1):
            ws_num.column_dimensions[get_column_letter(i)].width = w
            _hdr_cell(ws_num, 1, i, h, fill="2E75B6", size=9)

        for r, cp in enumerate(num_cols_prof, start=2):
            row_data = [
                cp.name, cp.mean, cp.std, cp.min, cp.p25, cp.p50,
                cp.p75, cp.p95, cp.max, cp.skewness, cp.kurtosis,
                cp.n_zeros, cp.n_negative, cp.n_outliers_iqr, f"{cp.pct_missing:.1f}%"
            ]
            for c, val in enumerate(row_data, 1):
                cell = _val_cell(ws_num, r, c, val)
                if c in range(2, 10) and isinstance(val, float):
                    cell.number_format = "#,##0.00"
                if c == 13 and isinstance(val, int) and val > 0:
                    cell.font = Font(name="Arial", size=9, color="C00000", bold=True)
                if c == 14 and isinstance(val, int):
                    pct = val / max(cp.n_total, 1) * 100
                    if pct > 5:
                        cell.font = Font(name="Arial", size=9, color="FF9900", bold=True)

        ws_num.freeze_panes = "B2"
        ws_num.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}1"

    # ── Sheet 3: Categorical Profiles ─────────────────────────────────────────
    ws_cat = wb.create_sheet("Categorical Profiles")
    cat_cols_prof = [cp for cp in ds.columns if cp.inferred_type in ("categorical", "text")]
    if cat_cols_prof:
        _hdr_cell(ws_cat, 1, 1, "Column", fill="2E75B6", size=9)
        _hdr_cell(ws_cat, 1, 2, "Unique Values", fill="2E75B6", size=9)
        _hdr_cell(ws_cat, 1, 3, "% Missing", fill="2E75B6", size=9)
        _hdr_cell(ws_cat, 1, 4, "Top Values (value: count — %)", fill="2E75B6", size=9)
        ws_cat.column_dimensions["A"].width = 20
        ws_cat.column_dimensions["B"].width = 15
        ws_cat.column_dimensions["C"].width = 12
        ws_cat.column_dimensions["D"].width = 70

        for r, cp in enumerate(cat_cols_prof, start=2):
            top_str = "  |  ".join(
                f"{v}: {c:,} ({p}%)" for v, c, p in cp.top_values[:8]
            ) if cp.top_values else "—"
            row_data = [cp.name, cp.n_unique, f"{cp.pct_missing:.1f}%", top_str]
            for c, val in enumerate(row_data, 1):
                cell = _val_cell(ws_cat, r, c, val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws_cat.row_dimensions[r].height = 28

        ws_cat.freeze_panes = "A2"

    # ── Sheet 4: Correlations ─────────────────────────────────────────────────
    if ds.correlations:
        ws_corr = wb.create_sheet("Correlations")
        cols = list(ds.correlations.keys())
        _hdr_cell(ws_corr, 1, 1, "Correlation Matrix (numeric columns)", size=10)
        ws_corr.merge_cells(f"A1:{get_column_letter(len(cols)+1)}1")

        for c, col in enumerate(cols, 2):
            _hdr_cell(ws_corr, 2, c, col, fill="2E75B6", size=8)
            ws_corr.column_dimensions[get_column_letter(c)].width = 12
        ws_corr.column_dimensions["A"].width = 20

        for r, row_col in enumerate(cols, start=3):
            _hdr_cell(ws_corr, r, 1, row_col, fill="2E75B6", size=8)
            for c, col_col in enumerate(cols, 2):
                val = ds.correlations.get(row_col, {}).get(col_col, 0)
                cell = ws_corr.cell(r, c, round(val, 3))
                cell.font = Font(name="Arial", size=8)
                cell.alignment = Alignment(horizontal="center")
                cell.border = _thin()
                abs_val = abs(val)
                if row_col == col_col:
                    cell.fill = PatternFill("solid", fgColor="D9D9D9")
                elif abs_val >= 0.9:
                    cell.fill = PatternFill("solid", fgColor="C00000")
                    cell.font = Font(name="Arial", size=8, color="FFFFFF", bold=True)
                elif abs_val >= 0.7:
                    cell.fill = PatternFill("solid", fgColor="FF9900")
                    cell.font = Font(name="Arial", size=8, bold=True)
                elif abs_val >= 0.5:
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")

        if ds.high_correlations:
            start = len(cols) + 5
            ws_corr.cell(start, 1, "High correlations (|r| ≥ 0.7)").font = Font(
                bold=True, name="Arial", size=10, color="1F4E79"
            )
            _hdr_cell(ws_corr, start + 1, 1, "Column A", fill="2E75B6", size=9)
            _hdr_cell(ws_corr, start + 1, 2, "Column B", fill="2E75B6", size=9)
            _hdr_cell(ws_corr, start + 1, 3, "Pearson r", fill="2E75B6", size=9)
            for i, (a, b, r_val) in enumerate(ds.high_correlations, start=start + 2):
                ws_corr.cell(i, 1, a).font = Font(name="Arial", size=9)
                ws_corr.cell(i, 2, b).font = Font(name="Arial", size=9)
                cell = ws_corr.cell(i, 3, r_val)
                cell.font = Font(name="Arial", size=9, bold=True,
                                 color="C00000" if abs(r_val) >= 0.9 else "FF9900")

    wb.save(path)
    print(f"  ✓ Excel → {path}")


# ── HTML EXPORT ───────────────────────────────────────────────────────────────

HTML_TMPL = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Data Profile — {{ ds.dataset_name }}</title>
<style>
  body{font-family:Arial,sans-serif;font-size:13px;color:#333;max-width:1200px;
       margin:40px auto;padding:0 24px;background:#f5f5f5}
  h1{color:#1F4E79;font-size:22px;margin-bottom:4px}
  h2{color:#2E75B6;font-size:15px;margin:32px 0 10px;border-bottom:2px solid #BDD7EE;padding-bottom:4px}
  h3{color:#1F4E79;font-size:13px;margin:18px 0 6px}
  .meta{color:#666;font-size:12px;margin-bottom:20px}
  .cards{display:flex;flex-wrap:wrap;gap:12px;margin-bottom:24px}
  .card{background:#fff;border:1px solid #ddd;border-radius:8px;padding:14px 18px;min-width:130px;text-align:center}
  .card .lbl{font-size:11px;color:#888;margin-bottom:4px}
  .card .val{font-size:20px;font-weight:bold;color:#1F4E79}
  .warn-box{background:#FFF4CC;border:1px solid #FFD966;border-radius:6px;
             padding:10px 16px;margin-bottom:20px}
  .warn-box li{color:#7F6000;font-size:12px;margin:3px 0}
  table{width:100%;border-collapse:collapse;background:#fff;border-radius:8px;
        margin-bottom:20px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,.08)}
  th{background:#1F4E79;color:#fff;padding:8px 12px;text-align:left;font-size:11px}
  td{padding:7px 12px;border-bottom:1px solid #eee;font-size:12px;vertical-align:top}
  tr:last-child td{border-bottom:none}
  tr:hover td{background:#f0f6ff}
  .badge{display:inline-block;padding:2px 8px;border-radius:10px;
         font-size:10px;font-weight:bold;color:#fff}
  .b-num{background:#2E75B6}.b-cat{background:#00B050}
  .b-dt{background:#FF9900}.b-txt{background:#7030A0}.b-bool{background:#666}
  .miss-hi{color:#C00000;font-weight:bold}
  .col-section{background:#fff;border:1px solid #ddd;border-radius:8px;
               padding:16px 20px;margin-bottom:16px}
  .col-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px}
  .stat-table td:first-child{color:#666;width:45%}
  .stat-table td:last-child{font-weight:bold}
  img{max-width:100%;border-radius:4px;margin-top:8px}
  .footer{color:#aaa;font-size:11px;margin-top:40px;text-align:center}
  details summary{cursor:pointer;color:#2E75B6;font-weight:bold;font-size:13px}
</style>
</head>
<body>
<h1>Data Profile Report</h1>
<div class="meta">
  Dataset: <b>{{ ds.dataset_name }}</b> &nbsp;|&nbsp;
  Source: <b>{{ ds.source_path }}</b> &nbsp;|&nbsp;
  Run: <b>{{ ds.run_timestamp }}</b>
</div>

<div class="cards">
  <div class="card"><div class="lbl">Rows</div><div class="val">{{ "{:,}".format(ds.n_rows) }}</div></div>
  <div class="card"><div class="lbl">Columns</div><div class="val">{{ ds.n_columns }}</div></div>
  <div class="card"><div class="lbl">Duplicates</div><div class="val">{{ ds.n_duplicates }}</div></div>
  <div class="card"><div class="lbl">Completeness</div><div class="val">{{ ds.overall_completeness }}%</div></div>
  <div class="card"><div class="lbl">Memory</div><div class="val">{{ ds.memory_mb }} MB</div></div>
</div>

{% if ds.warnings %}
<div class="warn-box">
  <b>⚠ Warnings</b>
  <ul>{% for w in ds.warnings %}<li>{{ w }}</li>{% endfor %}</ul>
</div>
{% endif %}

<h2>Column Overview</h2>
<table>
<thead><tr>
  <th>Column</th><th>Type</th><th>Inferred</th>
  <th>Missing</th><th>Unique</th><th>Summary</th>
</tr></thead>
<tbody>
{% for cp in ds.columns %}
<tr>
  <td><b>{{ cp.name }}</b></td>
  <td><code>{{ cp.dtype }}</code></td>
  <td>
    {% if cp.inferred_type == "numeric" %}<span class="badge b-num">numeric</span>
    {% elif cp.inferred_type == "categorical" %}<span class="badge b-cat">categorical</span>
    {% elif cp.inferred_type == "datetime" %}<span class="badge b-dt">datetime</span>
    {% elif cp.inferred_type == "text" %}<span class="badge b-txt">text</span>
    {% else %}<span class="badge b-bool">{{ cp.inferred_type }}</span>{% endif %}
  </td>
  <td class="{{ 'miss-hi' if cp.pct_missing > 10 else '' }}">
    {{ cp.pct_missing }}% ({{ "{:,}".format(cp.n_missing) }})
  </td>
  <td>{{ "{:,}".format(cp.n_unique) }} ({{ cp.pct_unique }}%)</td>
  <td>
    {% if cp.inferred_type == "numeric" and cp.mean is not none %}
      mean={{ "{:,.2f}".format(cp.mean) }}, std={{ "{:,.2f}".format(cp.std) }},
      IQR outliers={{ cp.n_outliers_iqr }}
    {% elif cp.top_values %}
      top: {{ cp.top_values[0][0] }} ({{ cp.top_values[0][2] }}%)
    {% elif cp.inferred_type == "datetime" %}
      {{ cp.date_min }} → {{ cp.date_max }}
    {% endif %}
  </td>
</tr>
{% endfor %}
</tbody>
</table>

<h2>Column Profiles</h2>
{% for cp in ds.columns %}
<details>
  <summary>{{ cp.name }} &nbsp;<small style="color:#888;font-weight:normal">
    {{ cp.inferred_type }} | {{ cp.pct_missing }}% missing | {{ "{:,}".format(cp.n_unique) }} unique
  </small></summary>
  <div class="col-section">
    <div class="col-grid">
      <div>
        <h3>Statistics</h3>
        <table class="stat-table">
          <tr><td>Total rows</td><td>{{ "{:,}".format(cp.n_total) }}</td></tr>
          <tr><td>Missing</td><td>{{ "{:,}".format(cp.n_missing) }} ({{ cp.pct_missing }}%)</td></tr>
          <tr><td>Unique values</td><td>{{ "{:,}".format(cp.n_unique) }} ({{ cp.pct_unique }}%)</td></tr>
          {% if cp.inferred_type == "numeric" and cp.mean is not none %}
          <tr><td>Mean</td><td>{{ "{:,.4f}".format(cp.mean) }}</td></tr>
          <tr><td>Std dev</td><td>{{ "{:,.4f}".format(cp.std) }}</td></tr>
          <tr><td>Min</td><td>{{ "{:,.4f}".format(cp.min) }}</td></tr>
          <tr><td>P25</td><td>{{ "{:,.4f}".format(cp.p25) }}</td></tr>
          <tr><td>Median</td><td>{{ "{:,.4f}".format(cp.p50) }}</td></tr>
          <tr><td>P75</td><td>{{ "{:,.4f}".format(cp.p75) }}</td></tr>
          <tr><td>P95</td><td>{{ "{:,.4f}".format(cp.p95) }}</td></tr>
          <tr><td>Max</td><td>{{ "{:,.4f}".format(cp.max) }}</td></tr>
          <tr><td>Skewness</td><td>{{ cp.skewness }}</td></tr>
          <tr><td>Kurtosis</td><td>{{ cp.kurtosis }}</td></tr>
          <tr><td>Zeros</td><td>{{ "{:,}".format(cp.n_zeros) }}</td></tr>
          <tr><td>Negative</td><td>{{ "{:,}".format(cp.n_negative) }}</td></tr>
          <tr><td>IQR outliers</td><td>{{ "{:,}".format(cp.n_outliers_iqr) }}</td></tr>
          {% elif cp.inferred_type == "datetime" %}
          <tr><td>Min date</td><td>{{ cp.date_min }}</td></tr>
          <tr><td>Max date</td><td>{{ cp.date_max }}</td></tr>
          <tr><td>Range (days)</td><td>{{ cp.date_range_days }}</td></tr>
          {% elif cp.inferred_type == "text" %}
          <tr><td>Avg length</td><td>{{ cp.avg_length }}</td></tr>
          <tr><td>Max length</td><td>{{ cp.max_length }}</td></tr>
          {% endif %}
        </table>
      </div>
      <div>
        {% if cp.chart_b64 %}
        <img src="data:image/png;base64,{{ cp.chart_b64 }}" alt="chart for {{ cp.name }}">
        {% endif %}
        {% if cp.top_values and cp.inferred_type in ["categorical", "text"] %}
        <h3>Top values</h3>
        <table>
          <tr><th>Value</th><th>Count</th><th>%</th></tr>
          {% for v, c, p in cp.top_values[:10] %}
          <tr><td>{{ v }}</td><td>{{ "{:,}".format(c) }}</td><td>{{ p }}%</td></tr>
          {% endfor %}
        </table>
        {% endif %}
      </div>
    </div>
  </div>
</details>
{% endfor %}

{% if ds.high_correlations %}
<h2>High Correlations (|r| ≥ 0.7)</h2>
<table>
  <thead><tr><th>Column A</th><th>Column B</th><th>Pearson r</th></tr></thead>
  <tbody>
  {% for a, b, r in ds.high_correlations %}
  <tr>
    <td>{{ a }}</td><td>{{ b }}</td>
    <td><b style="color:{{ '#C00000' if r|abs >= 0.9 else '#FF9900' }}">{{ r }}</b></td>
  </tr>
  {% endfor %}
  </tbody>
</table>
{% endif %}

<div class="footer">Generated automatically by data_profiler.py</div>
</body>
</html>
"""

def export_html(ds: DatasetProfile, path: str):
    from jinja2 import Template
    html = Template(HTML_TMPL).render(ds=ds)
    with open(path, "w") as f:
        f.write(html)
    print(f"  ✓ HTML  → {path}")


# ── MAIN ──────────────────────────────────────────────────────────────────────

def run(input_path: str, output_prefix: str, dataset_name: str = None):
    log = get_logger()
    t0  = time.time()

    df   = pd.read_csv(input_path)
    name = dataset_name or os.path.basename(input_path).replace(".csv", "")

    log_pipeline_start(log, "Data Profiling", input_path, len(df), len(df.columns))

    ds = profile_dataset(df, dataset_name=name, source_path=input_path)

    log.info(f"Overall completeness │  {ds.overall_completeness:.1f}%")
    log.info(f"Duplicate rows       │  {ds.n_duplicates:,} ({ds.pct_duplicates:.1f}%)")
    log.info(f"Memory               │  {ds.memory_mb:.2f} MB")

    log.info("Exporting outputs ...")
    os.makedirs(os.path.dirname(output_prefix) or ".", exist_ok=True)
    export_json(ds,  f"{output_prefix}.json")
    export_excel(ds, f"{output_prefix}.xlsx")
    export_html(ds,  f"{output_prefix}.html")

    log_pipeline_end(log, "Data Profiling", time.time() - t0, os.path.dirname(output_prefix) or ".")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Data Profiling Pipeline")
    parser.add_argument("--input",  default="transactions.csv")
    parser.add_argument("--output", default="profile_report")
    parser.add_argument("--name",   default=None)
    args = parser.parse_args()
    run(args.input, args.output, args.name)
